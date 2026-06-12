"""
app.py — Application web (Streamlit) d'économétrie financière
=============================================================
Donnez un fichier Excel contenant des séries HEBDOMADAIRES de prix (ou de
rendements) d'indices boursiers ; l'application réalise :

  1. Test de stationnarité ADF
  2. Test ARCH-LM (effets ARCH)
  3. GARCH univarié
  4. Causalité de Granger
  5. Test de contagion de Forbes-Rigobon

Lancement :  streamlit run app.py
"""
import io
import os

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st

import econometrics as ec

# ----------------------------------------------------------------------
# Configuration générale
# ----------------------------------------------------------------------
st.set_page_config(
    page_title="Économétrie financière — Indices boursiers",
    page_icon="📈",
    layout="wide",
)

plt.rcParams.update({
    "figure.facecolor": "white",
    "axes.grid": True,
    "grid.alpha": 0.25,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "font.size": 10,
})

BLEU = "#1f4e79"
ROUGE = "#c0392b"
ORANGE = "#e67e22"

# Masque l'icône de téléchargement CSV native des tableaux (elle produit un CSV
# virgule illisible sous Excel français). On garde la recherche et le plein écran.
# Les téléchargements passent par les boutons « Excel » dédiés, toujours propres.
st.markdown(
    """
    <style>
      [data-testid="stElementToolbar"] button[title="Download as CSV"] { display: none !important; }
    </style>
    """,
    unsafe_allow_html=True,
)


# ----------------------------------------------------------------------
# Fonctions utilitaires
# ----------------------------------------------------------------------
def fmt_p(p: float) -> str:
    """Formatage lisible d'une p-value."""
    if p < 1e-4:
        return f"{p:.2e}"
    return f"{p:.4f}"


def verdict(ok: bool, texte_ok: str, texte_non: str):
    """Affiche une conclusion en vert (ok) ou orange (non)."""
    if ok:
        st.success("✅ " + texte_ok)
    else:
        st.warning("⚠️ " + texte_non)


@st.cache_data(show_spinner=False)
def lire_fichier(contenu: bytes, nom: str, feuille):
    """Lit un fichier Excel/CSV en DataFrame (mis en cache)."""
    bio = io.BytesIO(contenu)
    if nom.lower().endswith(".csv"):
        # Détection simple du séparateur
        try:
            return pd.read_csv(bio, sep=None, engine="python")
        except Exception:
            bio.seek(0)
            return pd.read_csv(bio, sep=";")
    return pd.read_excel(bio, sheet_name=feuille)


@st.cache_data(show_spinner=False)
def noms_feuilles(contenu: bytes, nom: str):
    if nom.lower().endswith(".csv"):
        return None
    return pd.ExcelFile(io.BytesIO(contenu)).sheet_names


def arrondir_pour_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Arrondit toute valeur numérique à 3 décimales pour l'export Excel ;
    remplace par « ≈ 0 » ce qui s'afficherait 0,000 ou en notation scientifique
    (ex. 2,3e-308). Les entiers, booléens et textes restent intacts."""
    def fmt(v):
        if isinstance(v, (bool, np.bool_)):
            return bool(v)
        if isinstance(v, (int, np.integer)):
            return int(v)
        if isinstance(v, (float, np.floating)):
            if pd.isna(v):
                return v
            r = round(float(v), 3)
            return "≈ 0" if abs(r) < 0.0005 else r
        return v
    return df.apply(lambda col: col.map(fmt))


def telecharger_df(df: pd.DataFrame, label: str, fichier: str, key=None):
    # Export en vrai .xlsx : chaque variable dans sa propre colonne. Évite le
    # problème du CSV ouvert en colonne unique sous Excel français (séparateur
    # point-virgule). L'index n'est écrit que s'il porte de l'information
    # (dates, noms de séries) — pas un simple 0,1,2…
    if not fichier.lower().endswith(".xlsx"):
        fichier = fichier.rsplit(".", 1)[0] + ".xlsx"
    ecrire_index = not isinstance(df.index, pd.RangeIndex)
    out = df
    if ecrire_index and df.index.name is None:        # évite l'en-tête « Unnamed: 0 »
        out = df.copy()
        out.index.name = "Élément"
    out = arrondir_pour_excel(out)                    # max 3 décimales, « ≈ 0 » sinon
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        out.to_excel(writer, index=ecrire_index, sheet_name="Donnees")
        ws = writer.sheets["Donnees"]
        for col in ws.columns:
            largeur = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(largeur + 2, 40)
    st.download_button(
        label, buffer.getvalue(), file_name=fichier,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )


def ecrire_rapport_xlsx(feuilles: dict) -> bytes:
    """Écrit un dict {nom_feuille: DataFrame} dans un classeur Excel mis en forme
    (en-têtes bleus, volets figés, largeurs de colonnes ajustées)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    entete_font = Font(bold=True, color="FFFFFF")
    entete_fill = PatternFill("solid", fgColor="1F4E79")
    centre = Alignment(horizontal="center", vertical="center")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nom, df in feuilles.items():
            sname = str(nom)[:31]
            df = arrondir_pour_excel(df)              # max 3 décimales, « ≈ 0 » sinon
            ecrire_index = not isinstance(df.index, pd.RangeIndex)
            df.to_excel(writer, sheet_name=sname, index=ecrire_index)
            ws = writer.sheets[sname]
            for cell in ws[1]:
                cell.font = entete_font
                cell.fill = entete_fill
                cell.alignment = centre
            ws.freeze_panes = "A2"
            for col in ws.columns:
                largeur = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max(largeur + 2, 11), 44)
    return buffer.getvalue()


@st.cache_data(show_spinner=False)
def calc_vol_matrix(returns_sel: pd.DataFrame, dist: str):
    return ec.garch_vol_matrix(returns_sel, dist=dist)


@st.cache_data(show_spinner=False)
def calc_dy(vol_df: pd.DataFrame, lags: int, horizon: int):
    return ec.diebold_yilmaz(vol_df, lags=lags, horizon=horizon)


@st.cache_data(show_spinner=False)
def calc_dy_rolling(vol_df: pd.DataFrame, lags: int, horizon: int, window: int, step: int):
    return ec.dy_rolling_total(vol_df, lags=lags, horizon=horizon, window=window, step=step)


# ======================================================================
# BARRE LATÉRALE — chargement & préparation des données
# ======================================================================
st.sidebar.title("📂 Données")
st.sidebar.caption("Série(s) d'indices boursiers (hebdomadaires ou quotidiennes).")

fichier = st.sidebar.file_uploader(
    "Fichier Excel ou CSV",
    type=["xlsx", "xls", "csv"],
    help="1ʳᵉ colonne = dates ; colonnes suivantes = indices (prix ou rendements).",
)

EXEMPLE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exemple_donnees.xlsx")
MODELE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "modele_donnees.xlsx")


@st.dialog("📖 Guide : préparer votre fichier", width="large")
def guide_format():
    st.markdown(
        """
        Votre fichier doit suivre une structure **simple et régulière**. Voici le
        modèle à respecter :

        - **1ʳᵉ colonne = les dates** (une par ligne). Formats acceptés :
          `jj/mm/aaaa` (ex. `08/01/2010`) ou `aaaa-mm-jj`. L'ordre croissant ou
          décroissant n'a pas d'importance : l'application trie automatiquement.
        - **Colonnes suivantes = un indice par colonne**, avec son nom en en-tête
          (ex. `S&P 500`, `CAC 40`, `MASI (Maroc)`). Au moins **2 séries** pour les
          tests Granger, Forbes-Rigobon et DCC.
        - **Une ligne = une période** (données **hebdomadaires** ou **quotidiennes** —
          réglez la fréquence dans la barre latérale), sans trous au milieu.
        - Les valeurs peuvent être des **prix** (niveaux) **ou** des **rendements** ;
          vous l'indiquez ensuite dans « Nature des données ».
        """
    )
    st.markdown("**Voici un document type — exemple de feuille de prix :**")
    exemple = pd.DataFrame({
        "Date": ["08/01/2010", "15/01/2010", "22/01/2010", "29/01/2010"],
        "S&P 500": [1136.00, 1091.80, 1073.90, 1066.20],
        "CAC 40": [3954.38, 3820.78, 3739.46, 3563.76],
        "MASI (Maroc)": [10810.91, 10920.14, 10928.44, 10926.96],
    })
    st.dataframe(exemple, width="stretch", hide_index=True)

    st.info(
        "💡 Plusieurs feuilles dans le même classeur ? Pas de problème : vous "
        "choisissez la feuille à analyser dans le menu « Feuille Excel » à gauche."
    )
    st.warning(
        "⚠️ Si vous fournissez des **rendements en décimal** (ex. `0,012` pour 1,2 %), "
        "cochez « Multiplier par 100 » après avoir choisi « Rendements », "
        "pour que les résultats soient en pourcentage."
    )
    if os.path.exists(MODELE_PATH):
        with open(MODELE_PATH, "rb") as _f:
            st.download_button(
                "⬇️ Télécharger un fichier-modèle (.xlsx) à remplir",
                _f.read(),
                file_name="modele_donnees.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )


if st.sidebar.button("📖 Comment préparer le fichier ?", width="stretch"):
    guide_format()

if fichier is None and os.path.exists(EXEMPLE_PATH):
    if st.sidebar.button("📊 Charger les données d'exemple", width="stretch"):
        st.session_state["use_example"] = True

# Source des données : fichier téléversé, sinon exemple (si demandé)
use_example = fichier is None and st.session_state.get("use_example", False) \
    and os.path.exists(EXEMPLE_PATH)

if fichier is None and not use_example:
    st.sidebar.caption(
        "Pas de fichier ? Cliquez sur « Charger les données d'exemple », "
        "ou générez-le avec `python generer_exemple.py`."
    )
    st.title("📈 Économétrie financière des indices boursiers")
    st.markdown(
        """
        Cette application modélise, à partir d'une **série de prix ou de rendements**
        d'indices boursiers (**hebdomadaire ou quotidienne**), les tests suivants :

        | # | Test | Question posée |
        |---|------|----------------|
        | 1 | **ADF** | La série est-elle *stationnaire* ? |
        | 2 | **ARCH-LM** | Y a-t-il des *effets ARCH* (volatilité groupée) ? |
        | 3 | **GARCH univarié** | Comment modéliser la *volatilité conditionnelle* ? |
        | 4 | **Causalité de Granger** | Un indice en *précède*-t-il un autre ? |
        | 5 | **Forbes-Rigobon** | Y a-t-il *contagion* entre deux marchés ? |
        | 6 | **DCC-GARCH bivarié** | Comment évolue la *corrélation* dans le temps ? |
        | 7 | **Diebold-Yilmaz** | Qui *transmet* sa volatilité à qui ? |
        | 8 | **Portefeuille DCC** | Quel *gain de diversification* concret ? |

        ---
        #### Pour commencer
        1. Préparez un fichier **Excel/CSV** : 1ʳᵉ colonne = **dates**, colonnes
           suivantes = **indices** (au moins **2** pour Granger et Forbes-Rigobon).
        2. Chargez-le via le panneau de gauche **📂 Données**.
        3. Indiquez s'il s'agit de **prix** ou de **rendements**.
        4. Parcourez les onglets, configurez et lancez chaque test.

        > 💡 Astuce : un fichier de démonstration (`exemple_donnees.xlsx`,
        > indices *MASI, CAC40, SP500, SOURCE_US*) illustre tous les tests, dont
        > un épisode de **contagion**.
        """
    )
    st.stop()

# --- Lecture du fichier ------------------------------------------------
if use_example:
    with open(EXEMPLE_PATH, "rb") as _f:
        contenu = _f.read()
    nom_fichier = "exemple_donnees.xlsx"
    st.sidebar.success("📊 Données d'exemple chargées.")
else:
    contenu = fichier.getvalue()
    nom_fichier = fichier.name

feuilles = noms_feuilles(contenu, nom_fichier)
feuille = 0
if feuilles and len(feuilles) > 1:
    feuille = st.sidebar.selectbox("Feuille Excel", feuilles)

try:
    df_raw = lire_fichier(contenu, nom_fichier, feuille)
except Exception as e:
    st.sidebar.error(f"Lecture impossible : {e}")
    st.stop()

if df_raw.shape[1] < 2:
    st.sidebar.error("Le fichier doit contenir une colonne de dates et au moins une série.")
    st.stop()

# --- Colonne de dates --------------------------------------------------
colonnes = list(df_raw.columns)


def deviner_date(cols, df):
    for c in cols:
        if "date" in str(c).lower() or "time" in str(c).lower():
            return c
    for c in cols:
        if pd.api.types.is_datetime64_any_dtype(df[c]):
            return c
    return cols[0]


col_date = st.sidebar.selectbox(
    "Colonne des dates",
    colonnes,
    index=colonnes.index(deviner_date(colonnes, df_raw)),
)

df = df_raw.copy()
try:
    df[col_date] = pd.to_datetime(df[col_date], errors="coerce", dayfirst=True)
    if df[col_date].isna().all():
        raise ValueError
    df = df.dropna(subset=[col_date]).set_index(col_date).sort_index()
    index_dates = True
except Exception:
    st.sidebar.warning("Colonne de dates non reconnue : numérotation simple utilisée.")
    df = df_raw.copy()
    df.index = range(len(df))
    index_dates = False

# --- Colonnes numériques (séries) -------------------------------------
series_cols = [c for c in df.columns if c != col_date and pd.api.types.is_numeric_dtype(df[c])]
# Tentative de conversion des colonnes texte → numérique (virgules, espaces)
for c in df.columns:
    if c not in series_cols and c != col_date:
        conv = pd.to_numeric(
            df[c].astype(str).str.replace(" ", "").str.replace(",", "."),
            errors="coerce",
        )
        if conv.notna().mean() > 0.8:
            df[c] = conv
            series_cols.append(c)

if not series_cols:
    st.sidebar.error("Aucune colonne numérique exploitable trouvée.")
    st.stop()

df = df[series_cols].astype(float)

# --- Plage d'étude : filtre de dates global ---------------------------
if index_dates and len(df) > 2:
    st.sidebar.divider()
    dmin_g = df.index.min().date()
    dmax_g = df.index.max().date()
    plage_etude = st.sidebar.slider(
        "Plage d'étude",
        min_value=dmin_g, max_value=dmax_g,
        value=(dmin_g, dmax_g),
        format="YYYY-MM-DD",
        help="Restreint TOUTE l'analyse à cette sous-période (ex. depuis 2015), "
             "même si le fichier contient un historique plus long.",
    )
    d1g, d2g = pd.Timestamp(plage_etude[0]), pd.Timestamp(plage_etude[1])
    df = df[(df.index >= d1g) & (df.index <= d2g)]
    if df.shape[0] < 10:
        st.sidebar.error("Plage trop courte (moins de 10 observations). Élargissez-la.")
        st.stop()
    if (plage_etude[0] > dmin_g) or (plage_etude[1] < dmax_g):
        st.sidebar.caption(
            f"Analyse restreinte : {plage_etude[0]} → {plage_etude[1]} "
            f"({df.shape[0]} observations)."
        )

# Si la fenêtre de données change, on invalide les résultats DCC mémorisés
_sig = (str(df.index.min()), str(df.index.max()), int(df.shape[0]))
if st.session_state.get("data_sig") != _sig:
    st.session_state["data_sig"] = _sig
    st.session_state.pop("dcc_resultat", None)
    st.session_state.pop("dccm_resultat", None)
    st.session_state.pop("rapport_xlsx", None)

# --- Nature des données : prix ou rendements ? ------------------------
st.sidebar.divider()
mode = st.sidebar.radio(
    "Nature des données",
    ["Prix (niveaux)", "Rendements"],
    help="« Prix » : l'app calcule les rendements. « Rendements » : utilisés tels quels.",
)

if mode == "Prix (niveaux)":
    type_rdt = st.sidebar.selectbox("Type de rendement", ["Logarithmique", "Arithmétique"])
    if type_rdt == "Logarithmique":
        returns = pd.DataFrame({c: ec.to_log_returns(df[c], pct=True) for c in series_cols})
    else:
        returns = pd.DataFrame({c: ec.to_simple_returns(df[c], pct=True) for c in series_cols})
    returns = returns.dropna(how="all")
    prices = df
else:
    en_pct = st.sidebar.checkbox(
        "Multiplier par 100 (données en décimal, ex. 0,012)", value=False
    )
    returns = df * 100.0 if en_pct else df.copy()
    prices = None

# Fréquence des données → pilote l'annualisation, les fenêtres et les libellés
freq = st.sidebar.radio(
    "Fréquence des données", ["Hebdomadaire", "Quotidienne"], key="freq",
    help="Hebdomadaire : annualisation ×√52. Quotidienne : ×√252 (jours de bourse).",
)
if freq == "Quotidienne":
    periods_per_year = 252
    unite, unite_pl, unite_abbr = "jour", "jours", "j"
    fenetre_roll = 120                       # ≈ 6 mois de bourse
    dy_win_min, dy_win_max, dy_win_def, dy_win_step = 120, 750, 250, 10
    dy_h_max = 30
else:
    periods_per_year = 52
    unite, unite_pl, unite_abbr = "semaine", "semaines", "sem."
    fenetre_roll = 26                        # ≈ 6 mois
    dy_win_min, dy_win_max, dy_win_def, dy_win_step = 52, 260, 104, 4
    dy_h_max = 20

st.sidebar.divider()
st.sidebar.metric("Séries détectées", len(series_cols))
st.sidebar.metric("Observations (rendements)", int(returns.dropna(how="all").shape[0]))
if index_dates and len(returns) > 1:
    st.sidebar.caption(
        f"Du {returns.index.min().date()} au {returns.index.max().date()}"
    )

# Seuil de significativité global
alpha_signif = st.sidebar.select_slider(
    "Seuil de significativité", options=[0.01, 0.05, 0.10], value=0.05
)

# Signature — monogramme sobre dans le coin de la barre latérale
st.sidebar.markdown(
    f"""
    <div style="margin-top:1.4rem; padding-top:0.9rem;
                border-top:1px solid rgba(128,128,128,0.25);
                display:flex; align-items:center; gap:0.6rem; opacity:0.9;">
      <div style="width:36px; height:36px; border-radius:9px;
                  background:linear-gradient(135deg,{BLEU},#2c6ca3);
                  display:flex; align-items:center; justify-content:center;
                  color:#fff; font-family:Georgia,'Times New Roman',serif;
                  font-weight:700; font-size:0.9rem; letter-spacing:1px;
                  box-shadow:0 1px 4px rgba(0,0,0,0.28);">AE</div>
      <div style="line-height:1.15;">
        <div style="font-size:0.82rem; font-weight:600; letter-spacing:2px;
                    color:#9aa0a6;">ALI&nbsp;EB</div>
        <div style="font-size:0.64rem; letter-spacing:0.6px; color:#888;">
            Conception &amp; développement</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# ======================================================================
# CORPS — onglets
# ======================================================================
st.title("📈 Économétrie financière des indices boursiers")

onglets = st.tabs([
    "🗂️ Données & descriptif",
    "1️⃣ Stationnarité (ADF)",
    "2️⃣ Effets ARCH (ARCH-LM)",
    "3️⃣ GARCH univarié",
    "4️⃣ Causalité de Granger",
    "5️⃣ Contagion (Forbes-Rigobon)",
    "6️⃣ DCC-GARCH bivarié",
    "7️⃣ Spillover (Diebold-Yilmaz)",
    "8️⃣ Portefeuille (DCC)",
    "📑 Rapport Excel",
    "ℹ️ Méthodologie",
])

# ----------------------------------------------------------------------
# Onglet 0 — Données & statistiques descriptives
# ----------------------------------------------------------------------
with onglets[0]:
    st.subheader("Aperçu des données")
    c1, c2 = st.columns([2, 1])
    with c1:
        st.dataframe(df, width="stretch", height=360)
        st.caption(
            f"Tableau complet et défilable : {len(df)} lignes affichées "
            f"(toutes les observations chargées sont utilisées dans les calculs)."
        )
        telecharger_df(df, "⬇️ Télécharger les données (Excel)",
                       "donnees.xlsx", key="dl_donnees")
    with c2:
        st.metric("Nombre de séries", len(series_cols))
        st.metric("Nombre de points", len(df))
        st.write("**Séries :** " + ", ".join(series_cols))

    st.divider()
    st.subheader("Statistiques descriptives des rendements")
    desc = pd.DataFrame({c: ec.descriptive_stats(returns[c]) for c in series_cols}).T
    st.dataframe(
        desc.style.format({
            "Moyenne": "{:.4f}", "Médiane": "{:.4f}", "Écart-type": "{:.4f}",
            "Minimum": "{:.3f}", "Maximum": "{:.3f}",
            "Skewness (asymétrie)": "{:.3f}", "Kurtosis (normale = 3)": "{:.3f}",
            "Jarque-Bera (stat.)": "{:.2f}", "Jarque-Bera (p-value)": "{:.4f}",
        }),
        width="stretch",
    )
    st.caption(
        "Kurtosis > 3 = queues épaisses ; Jarque-Bera (p-value) < 5 % = rejet de "
        "la normalité — fréquent en finance et justifie une loi de Student pour le GARCH."
    )
    telecharger_df(desc, "⬇️ Télécharger les statistiques (Excel)",
                   "statistiques_descriptives.xlsx", key="dl_desc")

    st.divider()
    st.subheader("Graphiques")
    serie_plot = st.selectbox("Série à visualiser", series_cols, key="plot_serie")
    g1, g2 = st.columns(2)
    with g1:
        if prices is not None:
            fig, ax = plt.subplots(figsize=(6, 3.2))
            ax.plot(prices.index, prices[serie_plot], color=BLEU, lw=1)
            ax.set_title(f"{serie_plot} — Prix")
            st.pyplot(fig)
        else:
            st.info("Données fournies en rendements : pas de série de prix.")
    with g2:
        fig, ax = plt.subplots(figsize=(6, 3.2))
        ax.plot(returns.index, returns[serie_plot], color=ROUGE, lw=0.7)
        ax.set_title(f"{serie_plot} — Rendements (%)")
        st.pyplot(fig)
    st.caption(
        "La présence d'**amas de volatilité** (périodes calmes vs agitées) "
        "sur le graphique des rendements annonce des effets ARCH."
    )

# ----------------------------------------------------------------------
# Onglet 1 — ADF
# ----------------------------------------------------------------------
with onglets[1]:
    st.subheader("Test de Dickey-Fuller Augmenté (ADF)")
    st.markdown(
        "**H₀ :** racine unitaire (série **non stationnaire**) &nbsp;•&nbsp; "
        "**H₁ :** série **stationnaire**. &nbsp; p-value < seuil ⟹ on rejette H₀ ⟹ **stationnaire**."
    )

    c1, c2, c3 = st.columns(3)
    serie = c1.selectbox("Série", series_cols, key="adf_serie")
    cible_options = ["Rendements"] + (["Prix (niveau)"] if prices is not None else [])
    cible = c2.selectbox("Appliquer sur", cible_options, key="adf_cible")
    reg = c3.selectbox(
        "Spécification",
        ["c — constante", "ct — constante + tendance", "n — aucune"],
        index=1 if cible.startswith("Prix") else 0,
        key="adf_reg",
    )
    reg_code = {"c — constante": "c", "ct — constante + tendance": "ct", "n — aucune": "n"}[reg]

    serie_data = prices[serie] if cible.startswith("Prix") else returns[serie]

    if st.button("▶️ Lancer le test ADF", key="run_adf"):
        try:
            res = ec.adf_test(serie_data, regression=reg_code)
            m1, m2, m3 = st.columns(3)
            m1.metric("Statistique ADF", f"{res['statistique_adf']:.3f}")
            m2.metric("p-value", fmt_p(res["p_value"]))
            m3.metric("Retards retenus", res["retards_utilises"])

            tab_crit = pd.DataFrame(
                {"Valeur critique": res["valeurs_critiques"]}
            )
            st.write("**Valeurs critiques :**")
            st.dataframe(tab_crit.T, width="stretch")

            stationnaire = res["p_value"] < alpha_signif
            verdict(
                stationnaire,
                f"À {int(alpha_signif*100)} %, on rejette H₀ : **{serie}** ({cible.lower()}) "
                f"est **STATIONNAIRE** (p = {fmt_p(res['p_value'])}).",
                f"À {int(alpha_signif*100)} %, on ne rejette pas H₀ : **{serie}** ({cible.lower()}) "
                f"est **NON STATIONNAIRE** (p = {fmt_p(res['p_value'])}). "
                f"→ Différenciez la série (utilisez les rendements).",
            )
            if cible.startswith("Prix") and not stationnaire:
                st.info(
                    "C'est le résultat habituel : les **prix** sont non stationnaires (I(1)), "
                    "alors que les **rendements** le deviennent (I(0))."
                )
        except Exception as e:
            st.error(f"Erreur : {e}")

# ----------------------------------------------------------------------
# Onglet 2 — ARCH-LM
# ----------------------------------------------------------------------
with onglets[2]:
    st.subheader("Test ARCH-LM (multiplicateur de Lagrange d'Engle)")
    st.markdown(
        "**H₀ :** absence d'effets ARCH &nbsp;•&nbsp; **H₁ :** présence d'effets ARCH. "
        "p-value < seuil ⟹ **hétéroscédasticité conditionnelle** ⟹ un **GARCH est justifié**."
    )
    c1, c2 = st.columns(2)
    serie = c1.selectbox("Série (rendements)", series_cols, key="arch_serie")
    nlags = c2.slider("Nombre de retards", 1, 24, 5, key="arch_lags")

    if st.button("▶️ Lancer le test ARCH-LM", key="run_arch"):
        try:
            res = ec.arch_lm_test(returns[serie], nlags=nlags)
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Statistique LM", f"{res['lm_stat']:.2f}")
            m2.metric("p-value (LM)", fmt_p(res["lm_p_value"]))
            m3.metric("Statistique F", f"{res['f_stat']:.2f}")
            m4.metric("p-value (F)", fmt_p(res["f_p_value"]))

            verdict(
                res["lm_p_value"] < alpha_signif,
                f"Effets ARCH **présents** sur **{serie}** (p = {fmt_p(res['lm_p_value'])}). "
                f"La volatilité est groupée → passez à l'onglet **GARCH**.",
                f"**Aucun** effet ARCH détecté sur **{serie}** (p = {fmt_p(res['lm_p_value'])}). "
                f"Un modèle GARCH n'est pas nécessairement justifié.",
            )
        except Exception as e:
            st.error(f"Erreur : {e}")

# ----------------------------------------------------------------------
# Onglet 3 — GARCH univarié
# ----------------------------------------------------------------------
with onglets[3]:
    st.subheader("Modèle GARCH univarié")
    st.markdown(
        "Modélise la **variance conditionnelle** σ²ₜ. La **persistance** (α+β) proche "
        "de 1 indique des chocs de volatilité durables."
    )
    c1, c2, c3, c4 = st.columns(4)
    serie = c1.selectbox("Série (rendements)", series_cols, key="garch_serie")
    modele = c2.selectbox("Modèle", ["GARCH", "EGARCH", "GJR-GARCH (APARCH)"], key="garch_modele")
    p = c3.slider("p (ordre ARCH)", 1, 3, 1, key="garch_p")
    q = c4.slider("q (ordre GARCH)", 1, 3, 1, key="garch_q")
    c5, c6 = st.columns(2)
    moyenne = c5.selectbox("Équation de moyenne", ["Constant", "Zero", "AR"], key="garch_mean")
    loi = c6.selectbox(
        "Loi des innovations",
        ["normal — Normale", "t — Student", "skewt — Student asymétrique", "ged — GED"],
        index=1, key="garch_dist",
    )
    loi_code = loi.split(" ")[0]
    vol_code = {"GARCH": "GARCH", "EGARCH": "EGARCH", "GJR-GARCH (APARCH)": "APARCH"}[modele]
    o = 1 if modele != "GARCH" else 0

    if st.button("▶️ Estimer le GARCH", key="run_garch"):
        try:
            with st.spinner("Estimation en cours…"):
                g = ec.garch_fit(
                    returns[serie], p=p, q=q, o=o, mean=moyenne, dist=loi_code, vol=vol_code
                )
            st.caption(f"Spécification estimée : **{g['spec']}**")

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Persistance (α+β)", f"{g['persistence']:.4f}")
            m2.metric("Log-vraisemblance", f"{g['loglik']:.1f}")
            m3.metric("AIC", f"{g['aic']:.1f}")
            m4.metric("BIC", f"{g['bic']:.1f}")

            # Tableau des coefficients
            coef = pd.DataFrame({
                "Coefficient": g["params"],
                "Écart-type": g["std_errors"],
                "t-stat": g["tstats"],
                "p-value": g["pvalues"],
            })
            coef["Signif."] = coef["p-value"].apply(
                lambda x: "***" if x < 0.01 else ("**" if x < 0.05 else ("*" if x < 0.10 else ""))
            )
            st.write("**Coefficients estimés :**")
            st.dataframe(
                coef.style.format({
                    "Coefficient": "{:.4f}", "Écart-type": "{:.4f}",
                    "t-stat": "{:.3f}", "p-value": "{:.4f}",
                }),
                width="stretch",
            )
            st.caption("Significativité : *** 1 %, ** 5 %, * 10 %.")
            telecharger_df(coef, "⬇️ Télécharger les coefficients (Excel)",
                           f"garch_{serie}.xlsx", key="dl_garch")

            # Persistance
            if g["persistence"] >= 0.99:
                st.warning(
                    f"Persistance ≈ {g['persistence']:.3f} (très proche de 1) : chocs de "
                    "volatilité quasi permanents (proche d'un IGARCH)."
                )
            else:
                st.info(
                    f"Persistance = {g['persistence']:.3f} : la demi-vie d'un choc de "
                    f"volatilité est d'environ "
                    f"{np.log(0.5)/np.log(g['persistence']):.1f} {unite_pl}."
                )

            # Diagnostic des résidus
            ar = g["arch_residuals"]
            verdict(
                ar["lm_p_value"] >= alpha_signif,
                f"Résidus standardisés **sans effet ARCH résiduel** "
                f"(ARCH-LM p = {fmt_p(ar['lm_p_value'])}) → le modèle capte bien la volatilité.",
                f"Effets ARCH **résiduels** détectés (p = {fmt_p(ar['lm_p_value'])}) → "
                f"essayez d'augmenter p/q ou de changer de loi/modèle.",
            )

            # Graphique de la volatilité conditionnelle
            vol = g["conditional_volatility"]
            fig, ax = plt.subplots(figsize=(10, 3.6))
            ax.plot(returns[serie].dropna().index, returns[serie].dropna().values,
                    color="0.7", lw=0.6, label="Rendements")
            ax.plot(vol.index, vol.values, color=ROUGE, lw=1.4, label="Volatilité conditionnelle σₜ")
            ax.plot(vol.index, -vol.values, color=ROUGE, lw=1.4)
            ax.set_title(f"{serie} — Volatilité conditionnelle estimée")
            ax.legend(loc="upper left", fontsize=8)
            st.pyplot(fig)

            telecharger_df(
                vol.to_frame("volatilite_conditionnelle"),
                "⬇️ Télécharger la volatilité conditionnelle (Excel)",
                f"volatilite_{serie}.xlsx",
            )

            with st.expander("Voir le résumé complet du modèle (arch)"):
                st.text(g["summary"])
        except Exception as e:
            st.error(f"Erreur : {e}")

# ----------------------------------------------------------------------
# Onglet 4 — Granger
# ----------------------------------------------------------------------
with onglets[4]:
    st.subheader("Test de causalité de Granger")
    st.markdown(
        "**H₀ :** X ne cause pas Y (au sens de Granger). p-value < seuil ⟹ **X précède / "
        "aide à prévoir Y**. Les séries doivent être **stationnaires** (utilisez les rendements)."
    )
    if len(series_cols) < 2:
        st.warning("Il faut au moins **2 séries** pour ce test.")
    else:
        c1, c2, c3 = st.columns(3)
        serie_a = c1.selectbox("Série A", series_cols, index=0, key="gr_a")
        serie_b = c2.selectbox("Série B", series_cols,
                               index=1 if len(series_cols) > 1 else 0, key="gr_b")
        maxlag = c3.slider("Retards maximum", 1, 12, 4, key="gr_lags")

        if serie_a == serie_b:
            st.warning("Choisissez deux séries différentes.")
        elif st.button("▶️ Lancer le test de Granger", key="run_granger"):
            try:
                res = ec.granger_bidirectional(returns[serie_a], returns[serie_b], maxlag=maxlag)
                col1, col2 = st.columns(2)
                for col, sens, key in [
                    (col1, f"{serie_a} → {serie_b}", "a_vers_b"),
                    (col2, f"{serie_b} → {serie_a}", "b_vers_a"),
                ]:
                    with col:
                        st.markdown(f"**{sens}**")
                        tab = res[key]
                        st.dataframe(
                            tab.style.format({
                                "F": "{:.3f}", "p-value (F)": "{:.4f}",
                                "Chi²": "{:.3f}", "p-value (Chi²)": "{:.4f}",
                            }),
                            width="stretch", hide_index=True,
                        )
                        pmin = tab["p-value (F)"].min()
                        verdict(
                            pmin < alpha_signif,
                            f"**{sens.split('→')[0].strip()}** cause **{sens.split('→')[1].strip()}** "
                            f"(p min = {fmt_p(pmin)}).",
                            f"Pas de causalité {sens} (p min = {fmt_p(pmin)}).",
                        )
                st.caption(
                    "Causalité dans les **deux sens** = rétroaction (feedback). "
                    "Rappel : la causalité de Granger est une **précédence temporelle**, "
                    "pas une causalité au sens strict."
                )
            except Exception as e:
                st.error(f"Erreur : {e}")

# ----------------------------------------------------------------------
# Onglet 5 — Forbes-Rigobon
# ----------------------------------------------------------------------
with onglets[5]:
    st.subheader("Test de contagion de Forbes-Rigobon (2002)")
    st.markdown(
        "Pendant une crise, la corrélation entre marchés augmente mécaniquement avec la "
        "**volatilité** du marché-source. Forbes-Rigobon **corrige ce biais** pour distinguer "
        "la **contagion** d'une simple **interdépendance**."
    )
    st.markdown(
        "**H₀ :** ρ\\* (ajustée, crise) = ρ (stable) → *pas de contagion* &nbsp;•&nbsp; "
        "**H₁ :** ρ\\* > ρ (stable) → *contagion*."
    )
    if len(series_cols) < 2:
        st.warning("Il faut au moins **2 séries** pour ce test.")
    elif not index_dates:
        st.warning(
            "Ce test nécessite des **dates** valides pour délimiter la période de crise. "
            "Vérifiez la colonne des dates dans le panneau de gauche."
        )
    else:
        c1, c2 = st.columns(2)
        serie_src = c1.selectbox(
            "Marché-SOURCE (origine de la crise)", series_cols, index=len(series_cols) - 1,
            key="fr_src", help="Marché dont la hausse de volatilité sert à l'ajustement.",
        )
        serie_rcp = c2.selectbox(
            "Marché-RÉCEPTEUR (potentiellement contaminé)", series_cols, index=0, key="fr_rcp",
        )

        dmin = returns.index.min().date()
        dmax = returns.index.max().date()
        st.markdown("**Période de crise** (le reste = période stable) :")
        plage = st.slider(
            "Fenêtre de crise",
            min_value=dmin, max_value=dmax,
            value=(returns.index[int(len(returns) * 0.57)].date(),
                   returns.index[int(len(returns) * 0.66)].date()),
            format="YYYY-MM-DD", key="fr_plage",
        )
        bilateral = st.checkbox("Test bilatéral (par défaut : unilatéral H₁ : ρ\\* > ρ)", value=False)

        if serie_src == serie_rcp:
            st.warning("Choisissez deux marchés différents.")
        elif st.button("▶️ Lancer le test de Forbes-Rigobon", key="run_fr"):
            try:
                d1, d2 = pd.Timestamp(plage[0]), pd.Timestamp(plage[1])
                crisis_mask = pd.Series(
                    (returns.index >= d1) & (returns.index <= d2), index=returns.index
                )
                res = ec.forbes_rigobon_test(
                    returns[serie_src], returns[serie_rcp], crisis_mask,
                    one_sided=not bilateral,
                )

                m1, m2, m3 = st.columns(3)
                m1.metric("ρ — période stable", f"{res['rho_stable']:.3f}",
                          help=f"{res['n_stable']} observations")
                m2.metric("ρ — crise (brute)", f"{res['rho_crisis_brut']:.3f}",
                          f"{res['rho_crisis_brut'] - res['rho_stable']:+.3f}",
                          help=f"{res['n_crisis']} observations")
                m3.metric("ρ\\* — crise (ajustée)", f"{res['rho_crisis_ajuste']:.3f}",
                          f"{res['rho_crisis_ajuste'] - res['rho_stable']:+.3f}")

                m4, m5, m6 = st.columns(3)
                m4.metric("δ — hausse de variance", f"{res['delta_volatilite']:.2f}",
                          help="Var(crise)/Var(stable) − 1 du marché-source")
                m5.metric("Statistique z (ajustée)", f"{res['z_ajuste']:.3f}")
                m6.metric("p-value (ajustée)", fmt_p(res["p_value_ajuste"]))

                if res["delta_volatilite"] <= 0:
                    st.warning(
                        f"⚠️ **Test non valide dans ce sens : δ = {res['delta_volatilite']:.2f} ≤ 0.** "
                        f"La volatilité de la source (**{serie_src}**) n'a **pas augmenté** pendant "
                        f"la fenêtre de crise. Or Forbes-Rigobon suppose une **hausse** de la "
                        f"volatilité de la source : avec δ ≤ 0 l'ajustement **gonfle** la corrélation "
                        f"au lieu de la corriger, ce qui produit une **contagion factice**. "
                        f"→ Inversez le sens (prenez **{serie_rcp}** comme source) ou choisissez une "
                        f"fenêtre de crise où **{serie_src}** est réellement plus volatil."
                    )
                else:
                    verdict(
                        res["p_value_ajuste"] < alpha_signif,
                        f"**CONTAGION** de **{serie_src}** vers **{serie_rcp}** "
                        f"(p = {fmt_p(res['p_value_ajuste'])}). La corrélation augmente **au-delà** "
                        f"de ce qu'explique la hausse de volatilité.",
                        f"**PAS de contagion** ({serie_src} → {serie_rcp}, "
                        f"p = {fmt_p(res['p_value_ajuste'])}). La hausse de corrélation s'explique "
                        f"par la volatilité : **simple interdépendance** (« No Contagion, Only "
                        f"Interdependence »).",
                    )

                # Comparaison test brut vs ajusté
                st.markdown("**Effet de l'ajustement d'hétéroscédasticité :**")
                comp = pd.DataFrame({
                    "Corrélation": [res["rho_crisis_brut"], res["rho_crisis_ajuste"]],
                    "Stat. z": [res["z_brut"], res["z_ajuste"]],
                    "p-value": [res["p_value_brut"], res["p_value_ajuste"]],
                    "Contagion ?": [
                        "Oui" if res["p_value_brut"] < alpha_signif else "Non",
                        "Oui" if res["p_value_ajuste"] < alpha_signif else "Non",
                    ],
                }, index=["Test BRUT (non ajusté)", "Test AJUSTÉ (Forbes-Rigobon)"])
                st.dataframe(
                    comp.style.format({"Corrélation": "{:.3f}", "Stat. z": "{:.3f}", "p-value": "{:.4f}"}),
                    width="stretch",
                )
                telecharger_df(comp, "⬇️ Télécharger le test (Excel)",
                               f"forbes_rigobon_{serie_src}_{serie_rcp}.xlsx", key="dl_fr")
                if (res["p_value_brut"] < alpha_signif) and (res["p_value_ajuste"] >= alpha_signif):
                    st.info(
                        "💡 Cas emblématique de Forbes-Rigobon : le test **brut** conclut à tort "
                        "à une contagion, que l'**ajustement** fait disparaître."
                    )

                # Corrélation glissante + zone de crise
                roll = (
                    returns[serie_src].rolling(fenetre_roll).corr(returns[serie_rcp])
                )
                fig, ax = plt.subplots(figsize=(10, 3.4))
                ax.plot(roll.index, roll.values, color=BLEU, lw=1.2,
                        label=f"Corrélation glissante ({fenetre_roll} {unite_abbr})")
                ax.axvspan(d1, d2, color=ORANGE, alpha=0.25, label="Période de crise")
                ax.axhline(res["rho_stable"], color="0.5", ls="--", lw=1,
                           label=f"ρ stable = {res['rho_stable']:.2f}")
                ax.set_title(f"Corrélation {serie_src} – {serie_rcp}")
                ax.legend(loc="upper left", fontsize=8)
                st.pyplot(fig)
            except Exception as e:
                st.error(f"Erreur : {e}")

# ----------------------------------------------------------------------
# Onglet 6 — DCC-GARCH bivarié
# ----------------------------------------------------------------------
with onglets[6]:
    st.subheader("Modèle DCC-GARCH bivarié (Engle, 2002)")
    st.markdown(
        "Estime la **corrélation conditionnelle dynamique** ρₜ entre deux marchés, "
        "semaine par semaine. Méthode en deux étapes : un **GARCH univarié** sur chaque "
        "série, puis la dynamique de corrélation de paramètres **a** (réaction) et "
        "**b** (persistance), avec **a + b < 1**."
    )
    if len(series_cols) < 2:
        st.warning("Il faut au moins **2 séries** pour ce test.")
    else:
        c1, c2 = st.columns(2)
        serie_1 = c1.selectbox("Série 1", series_cols, index=0, key="dcc_a")
        serie_2 = c2.selectbox("Série 2", series_cols,
                               index=1 if len(series_cols) > 1 else 0, key="dcc_b")
        c3, c4 = st.columns(2)
        modele_dcc = c3.selectbox("Modèle de marge (GARCH)",
                                  ["GARCH", "EGARCH", "GJR-GARCH (APARCH)"], key="dcc_modele")
        loi_dcc = c4.selectbox(
            "Loi des innovations",
            ["t — Student", "normal — Normale", "skewt — Student asymétrique", "ged — GED"],
            index=0, key="dcc_dist",
        )
        loi_dcc_code = loi_dcc.split(" ")[0]
        vol_dcc_code = {"GARCH": "GARCH", "EGARCH": "EGARCH",
                        "GJR-GARCH (APARCH)": "APARCH"}[modele_dcc]
        o_dcc = 1 if modele_dcc != "GARCH" else 0

        lancer_dcc = st.button("▶️ Estimer le DCC-GARCH", key="run_dcc")
        if serie_1 == serie_2:
            st.warning("Choisissez deux séries différentes.")
        else:
            # --- estimation (uniquement au clic) : on mémorise le résultat ----
            if lancer_dcc:
                try:
                    with st.spinner("Estimation des deux GARCH puis du DCC…"):
                        d = ec.dcc_garch_bivariate(
                            returns[serie_1], returns[serie_2],
                            dist=loi_dcc_code, vol=vol_dcc_code, o=o_dcc,
                        )
                    st.session_state["dcc_resultat"] = d
                    st.session_state["dcc_paire"] = (serie_1, serie_2)
                except Exception as e:
                    st.session_state.pop("dcc_resultat", None)
                    st.error(f"Erreur : {e}")

            # --- affichage (persiste tant qu'un résultat existe) --------------
            d = st.session_state.get("dcc_resultat")
            if d is not None:
                p1, p2 = st.session_state.get("dcc_paire", (serie_1, serie_2))
                if (p1, p2) != (serie_1, serie_2):
                    st.info(
                        f"Résultat affiché pour **{p1} – {p2}**. Cliquez sur "
                        "« Estimer le DCC-GARCH » pour recalculer sur la nouvelle paire."
                    )
                st.caption(f"Spécification estimée : **{d['spec']}** — {d['n_obs']} observations.")

                m1, m2, m3, m4 = st.columns(4)
                m1.metric("a (réaction)", f"{d['a']:.4f}")
                m2.metric("b (persistance)", f"{d['b']:.4f}")
                m3.metric("a + b", f"{d['persistance']:.4f}")
                m4.metric("ρ moyenne", f"{d['rho_moyenne']:.3f}")

                # Tableau des paramètres DCC
                lignes = []
                for nom, est, se, pv in [
                    ("a (réaction aux chocs)", d["a"], d["a_se"], d["a_p"]),
                    ("b (persistance)", d["b"], d["b_se"], d["b_p"]),
                ]:
                    lignes.append({
                        "Paramètre": nom,
                        "Estimation": est,
                        "Écart-type": se if se is not None else np.nan,
                        "p-value": pv if pv is not None else np.nan,
                    })
                tab_dcc = pd.DataFrame(lignes)
                st.write("**Paramètres de la dynamique de corrélation :**")
                st.dataframe(
                    tab_dcc.style.format({
                        "Estimation": "{:.4f}", "Écart-type": "{:.4f}", "p-value": "{:.4f}",
                    }, na_rep="—"),
                    width="stretch", hide_index=True,
                )
                st.caption("Écarts-types numériques (Hessienne par différences finies), donc approximatifs.")
                telecharger_df(tab_dcc, "⬇️ Télécharger les paramètres DCC (Excel)",
                               f"dcc_parametres_{p1}_{p2}.xlsx", key="dl_dccparam")

                if d["persistance"] >= 0.999:
                    st.warning(
                        f"a + b ≈ {d['persistance']:.3f} (très proche de 1) : corrélation "
                        "conditionnelle quasi intégrée (chocs de corrélation très durables)."
                    )
                else:
                    st.info(
                        f"a + b = {d['persistance']:.3f} < 1 : la dynamique de corrélation est "
                        f"**stable** et revient vers sa moyenne de long terme "
                        f"(ρ̄ inconditionnelle ≈ {d['rho_inconditionnelle']:.3f})."
                    )

                # Statistiques sur la corrélation conditionnelle
                rho_t = d["rho_t"]
                rho_plot = rho_t            # par défaut : toute la plage
                titre_plage = "toute la période"
                s1, s2, s3, s4 = st.columns(4)
                s1.metric("ρ minimum", f"{d['rho_min']:.3f}")
                s2.metric("ρ maximum", f"{d['rho_max']:.3f}")
                s3.metric("ρ écart-type", f"{d['rho_ecart_type']:.3f}")
                s4.metric("Log-vraisemblance (DCC)", f"{d['loglik']:.1f}")

                # Moyennes par sous-période (si dates disponibles)
                if index_dates:
                    st.markdown("**Corrélation conditionnelle moyenne par sous-période :**")
                    dr1, dr2 = st.columns(2)
                    deb = dr1.date_input("Début sous-période", value=rho_t.index.min().date(),
                                         min_value=rho_t.index.min().date(),
                                         max_value=rho_t.index.max().date(), key="dcc_deb")
                    fin = dr2.date_input("Fin sous-période", value=rho_t.index.max().date(),
                                         min_value=rho_t.index.min().date(),
                                         max_value=rho_t.index.max().date(), key="dcc_fin")
                    masque = (rho_t.index >= pd.Timestamp(deb)) & (rho_t.index <= pd.Timestamp(fin))
                    if masque.sum() > 0:
                        st.metric(
                            f"ρ moyenne sur la fenêtre ({int(masque.sum())} obs.)",
                            f"{rho_t[masque].mean():.3f}",
                            f"{rho_t[masque].mean() - d['rho_moyenne']:+.3f} vs moyenne totale",
                        )
                        rho_plot = rho_t[masque]   # le graphe suit la sous-période
                        titre_plage = f"{deb} → {fin}"

                # Graphique de la corrélation conditionnelle (limité à la sous-période)
                fig, ax = plt.subplots(figsize=(10, 3.6))
                ax.plot(rho_plot.index, rho_plot.values, color=BLEU, lw=1.2,
                        label="Corrélation conditionnelle ρₜ (DCC)")
                ax.axhline(d["rho_inconditionnelle"], color="0.5", ls="--", lw=1,
                           label=f"ρ inconditionnelle = {d['rho_inconditionnelle']:.2f}")
                ax.set_title(f"DCC-GARCH : {p1} – {p2}  ({titre_plage})")
                ax.legend(loc="upper left", fontsize=8)
                st.pyplot(fig)
                st.caption(
                    "Le graphique et la moyenne suivent la **sous-période** choisie ci-dessus. "
                    "Pour revoir toute la plage, remettez les dates de début et de fin aux extrêmes."
                )

                telecharger_df(
                    rho_t.to_frame("rho_conditionnelle"),
                    "⬇️ Télécharger la corrélation conditionnelle (Excel)",
                    f"dcc_{p1}_{p2}.xlsx",
                )

        # ==============================================================
        # Version multivariée : N séries (matrice de corrélation complète)
        # ==============================================================
        st.divider()
        st.markdown("### 🔗 Version multivariée — plusieurs séries à la fois")
        st.markdown(
            "Même modèle, généralisé à **N marchés** : on obtient la **matrice de "
            "corrélation conditionnelle** complète et un **indice d'intégration** "
            "(corrélation moyenne entre tous les couples), semaine par semaine. "
            "Les paramètres **a** et **b** sont communs à toutes les paires (DCC scalaire)."
        )
        defaut_m = list(series_cols[:min(3, len(series_cols))])
        choix_m = st.multiselect(
            "Séries à inclure (2 minimum, idéalement 3 à 6)",
            list(series_cols), default=defaut_m, key="dccm_series",
        )
        cm1, cm2 = st.columns(2)
        modele_dccm = cm1.selectbox("Modèle de marge (GARCH)",
                                    ["GARCH", "EGARCH", "GJR-GARCH (APARCH)"], key="dccm_modele")
        loi_dccm = cm2.selectbox(
            "Loi des innovations",
            ["t — Student", "normal — Normale", "skewt — Student asymétrique", "ged — GED"],
            index=0, key="dccm_dist",
        )
        loi_dccm_code = loi_dccm.split(" ")[0]
        vol_dccm_code = {"GARCH": "GARCH", "EGARCH": "EGARCH",
                         "GJR-GARCH (APARCH)": "APARCH"}[modele_dccm]
        o_dccm = 1 if modele_dccm != "GARCH" else 0

        lancer_dccm = st.button("▶️ Estimer le DCC-GARCH multivarié", key="run_dccm")
        if len(choix_m) < 2:
            st.warning("Sélectionnez au moins **deux séries**.")
        else:
            if lancer_dccm:
                try:
                    with st.spinner(f"Estimation de {len(choix_m)} GARCH puis du DCC multivarié…"):
                        dm = ec.dcc_garch_multivariate(
                            [returns[c] for c in choix_m], noms=list(choix_m),
                            dist=loi_dccm_code, vol=vol_dccm_code, o=o_dccm,
                        )
                    st.session_state["dccm_resultat"] = dm
                    st.session_state["dccm_series_ok"] = list(choix_m)
                except Exception as e:
                    st.session_state.pop("dccm_resultat", None)
                    st.error(f"Erreur : {e}")

            dm = st.session_state.get("dccm_resultat")
            if dm is not None:
                noms_ok = st.session_state.get("dccm_series_ok", list(choix_m))
                if list(noms_ok) != list(choix_m):
                    st.info(
                        f"Résultat affiché pour : **{', '.join(noms_ok)}**. "
                        "Cliquez sur « Estimer » pour recalculer sur la nouvelle sélection."
                    )
                st.caption(f"Spécification : **{dm['spec']}** — {dm['n_obs']} observations.")

                mm1, mm2, mm3, mm4 = st.columns(4)
                mm1.metric("a (réaction)", f"{dm['a']:.4f}")
                mm2.metric("b (persistance)", f"{dm['b']:.4f}")
                mm3.metric("a + b", f"{dm['persistance']:.4f}")
                mm4.metric("Corr. moyenne", f"{dm['corr_moyenne']:.3f}")

                if dm["persistance"] >= 0.999:
                    st.warning(f"a + b ≈ {dm['persistance']:.3f} : corrélations quasi intégrées.")
                else:
                    st.info(f"a + b = {dm['persistance']:.3f} < 1 : dynamique de corrélation stable.")

                st.markdown("**Matrice de corrélation conditionnelle moyenne :**")
                st.dataframe(
                    dm["R_moyenne"].style.format("{:.3f}").background_gradient(
                        cmap="RdBu_r", vmin=-1, vmax=1),
                    width="stretch",
                )

                # indice d'intégration des marchés dans le temps
                integ = dm["corr_moyenne_t"]
                fig, ax = plt.subplots(figsize=(10, 3.4))
                ax.plot(integ.index, integ.values, color=BLEU, lw=1.2,
                        label="Corrélation moyenne entre tous les couples")
                ax.axhline(dm["corr_moyenne"], color="0.5", ls="--", lw=1,
                           label=f"moyenne = {dm['corr_moyenne']:.2f}")
                ax.set_title("Indice d'intégration des marchés (DCC multivarié)")
                ax.set_ylabel("ρ moyen"); ax.legend(loc="upper left", fontsize=8)
                st.pyplot(fig)
                st.caption(
                    "Quand cette courbe monte, les marchés bougent **plus ensemble** "
                    "(intégration / contagion) ; quand elle baisse, ils se **diversifient**."
                )

                # corrélation conditionnelle d'un couple précis
                paires = dm["paires"]
                if paires:
                    labels = [f"{i} ↔ {j}" for (i, j) in paires]
                    choix_lbl = st.selectbox("Voir la corrélation d'un couple précis",
                                             labels, key="dccm_pair")
                    i_sel, j_sel = paires[labels.index(choix_lbl)]
                    sc = dm["rho_pairs"][(i_sel, j_sel)]
                    fig, ax = plt.subplots(figsize=(10, 3.2))
                    ax.plot(sc.index, sc.values, color=ROUGE, lw=1.1,
                            label=f"ρₜ {i_sel} ↔ {j_sel}")
                    ax.axhline(float(sc.mean()), color="0.5", ls="--", lw=1,
                               label=f"moyenne = {sc.mean():.2f}")
                    ax.set_title(f"Corrélation conditionnelle : {choix_lbl}")
                    ax.legend(loc="upper left", fontsize=8)
                    st.pyplot(fig)

                telecharger_df(
                    dm["rho_pairs_df"],
                    "⬇️ Télécharger les corrélations par couple (Excel)",
                    "dcc_multivarie.xlsx",
                )

# ----------------------------------------------------------------------
# Onglet 7 — Spillover de volatilité (Diebold-Yilmaz)
# ----------------------------------------------------------------------
with onglets[7]:
    st.subheader("Spillover de volatilité — Diebold & Yilmaz (2012)")
    st.markdown(
        "Mesure **qui transmet sa volatilité à qui**. À partir d'un VAR sur les "
        "volatilités conditionnelles (GARCH), on décompose la variance d'erreur de "
        "prévision : l'**indice total** résume la connectivité du système, et le "
        "**spillover net** dit si un marché est plutôt **émetteur** (+) ou **récepteur** (−)."
    )
    if len(series_cols) < 2:
        st.warning("Il faut au moins **2 séries** pour ce test.")
    else:
        sel = st.multiselect("Séries à inclure", series_cols, default=series_cols, key="dy_sel")
        c1, c2, c3, c4 = st.columns(4)
        dy_lags = c1.slider("Retards du VAR", 1, 6, 2, key="dy_lags")
        dy_h = c2.slider(f"Horizon de prévision ({unite_abbr})", 4, dy_h_max, 10, key="dy_h")
        dy_win = c3.slider(f"Fenêtre glissante ({unite_abbr})", dy_win_min, dy_win_max,
                           dy_win_def, step=dy_win_step, key="dy_win")
        dy_dist = c4.selectbox("Loi des marges GARCH",
                               ["t — Student", "normal — Normale", "ged — GED"], key="dy_dist")
        dy_dist_code = dy_dist.split(" ")[0]

        if len(sel) < 2:
            st.warning("Sélectionnez au moins 2 séries.")
        elif st.button("▶️ Calculer les spillovers", key="run_dy"):
            try:
                with st.spinner("Estimation des GARCH puis du VAR…"):
                    vol_df = calc_vol_matrix(returns[sel].dropna(), dy_dist_code)
                    dy = calc_dy(vol_df, dy_lags, dy_h)

                # Garde-fou : un VAR estime k·(k·p+1) coefficients. Sur un
                # échantillon court il est sur-paramétré → décomposition instable.
                n_dy = vol_df.shape[0]
                par_eq = len(sel) * dy_lags + 1
                ratio_dy = n_dy / par_eq
                if ratio_dy < 10:
                    st.warning(
                        f"⚠️ **Échantillon trop court pour un spillover fiable.** Le VAR estime "
                        f"**{par_eq} coefficients par équation** à partir de seulement "
                        f"**{n_dy} observations** (≈ {ratio_dy:.1f} obs/coefficient). En dessous "
                        f"de ~10, la décomposition devient **très instable** : décaler la période "
                        f"de quelques {unite_pl} peut faire varier les résultats du tout au tout. "
                        f"→ Élargissez la **Plage d'étude**, réduisez le **nombre de séries**, "
                        f"ou baissez les **retards du VAR**."
                    )

                st.metric("📊 Indice de spillover TOTAL",
                          f"{dy['total']:.1f} %",
                          help="Part de la variance du système expliquée par les transmissions "
                               "entre marchés (le reste = chocs propres).")

                st.markdown("**Table de connectivité** (ligne = marché qui *reçoit*, "
                            "colonne = marché *source*, en %) :")
                tbl = dy["table"]
                st.dataframe(
                    tbl.style.format("{:.1f}", na_rep="—")
                    .background_gradient(cmap="Oranges", axis=None),
                    width="stretch",
                )
                st.caption("Diagonale = part de volatilité « propre » ; hors-diagonale = spillovers ; "
                           "case en bas à droite = indice total du système.")

                # Spillover net directionnel
                st.markdown("**Spillover net** (émis − reçu) : positif = émetteur, négatif = récepteur.")
                net = dy["net"].sort_values()
                fig, ax = plt.subplots(figsize=(9, 0.5 + 0.5 * len(net)))
                couleurs = [ROUGE if v < 0 else "#2e7d32" for v in net.values]
                ax.barh(net.index, net.values, color=couleurs)
                ax.axvline(0, color="0.4", lw=0.8)
                ax.set_xlabel("Spillover net (%)")
                ax.set_title("Émetteurs (+) vs récepteurs (−) de volatilité")
                st.pyplot(fig)

                tneg = net.idxmin()
                st.info(
                    f"**{tneg}** est le plus **net récepteur** de volatilité "
                    f"({net.min():+.1f} %) : cohérent avec un marché *price-taker*. "
                    f"À l'inverse, **{net.idxmax()}** est le plus net **émetteur** ({net.max():+.1f} %)."
                )

                # Indice total glissant
                with st.spinner("Calcul de l'indice glissant…"):
                    roll = calc_dy_rolling(vol_df, dy_lags, dy_h, dy_win, 2)
                if len(roll) > 2:
                    fig, ax = plt.subplots(figsize=(10, 3.4))
                    ax.plot(roll.index, roll.values, color=BLEU, lw=1.2)
                    ax.axhline(dy["total"], color="0.5", ls="--", lw=1,
                               label=f"Moyenne plein échantillon = {dy['total']:.0f} %")
                    ax.set_title(f"Indice de spillover total glissant ({dy_win} {unite_abbr})")
                    ax.set_ylabel("%"); ax.legend(loc="upper left", fontsize=8)
                    st.pyplot(fig)
                    st.caption("Les pics correspondent aux périodes de crise (intégration accrue des marchés).")

                telecharger_df(tbl, "⬇️ Télécharger la table de spillover (Excel)", "spillover_dy.xlsx")
            except Exception as e:
                st.error(f"Erreur : {e}")

# ----------------------------------------------------------------------
# Onglet 8 — Portefeuille à variance minimale (covariance DCC)
# ----------------------------------------------------------------------
with onglets[8]:
    st.subheader("Portefeuille à variance minimale dynamique (covariance DCC)")
    st.markdown(
        "Exploite la **covariance conditionnelle DCC** pour construire un portefeuille à "
        "**risque minimal** dont les poids **varient dans le temps**. Chiffre concrètement "
        "le **gain de diversification** et fournit un **ratio de couverture** dynamique."
    )
    st.warning(
        "⚠️ **Devises** : si les actifs ne sont pas dans la **même devise** "
        "(ex. S&P 500 en USD, CAC 40 en EUR, MASI en MAD), la volatilité annualisée "
        "ci-dessous **ignore le risque de change**. Elle correspond à un investisseur "
        "raisonnant en devise locale (ou parfaitement couvert contre le change). Pour une "
        "allocation réellement réalisable en une devise unique, convertissez d'abord les "
        "indices dans cette devise. *(Sans effet sur les onglets DCC / spillover, qui mesurent "
        "la co-évolution des marchés en devise locale — l'approche standard.)*"
    )
    d = st.session_state.get("dcc_resultat")
    dm = st.session_state.get("dccm_resultat")

    sources = []
    if d is not None:
        sources.append("Bivarié (2 actifs)")
    if dm is not None:
        sources.append("Multivarié (N actifs)")

    if not sources:
        st.warning(
            "Estimez d'abord un **DCC-GARCH** (onglet 6) — **bivarié** ou **multivarié**. "
            "Le portefeuille se construit sur les séries de ce modèle."
        )
    else:
        if len(sources) > 1:
            choix_src = st.radio("Construire le portefeuille à partir du DCC :",
                                 sources, horizontal=True, key="pf_src")
        else:
            choix_src = sources[0]
            st.caption(f"Source : **{choix_src}** (le seul DCC estimé pour l'instant).")
        long_only = st.checkbox("Interdire la vente à découvert (poids ≥ 0)",
                                value=True, key="pf_longonly")

        # ==============================================================
        # Cas A — portefeuille bivarié (2 actifs)
        # ==============================================================
        if choix_src.startswith("Bivarié"):
            p1, p2 = st.session_state.get("dcc_paire", (series_cols[0], series_cols[1]))
            st.caption(f"Paire issue du DCC : **{p1}** et **{p2}** — {d['n_obs']} {unite_pl}.")
            try:
                pf = ec.dcc_min_variance_portfolio(
                    d, returns[p1], returns[p2], allow_short=not long_only,
                    periods_per_year=periods_per_year,
                )
                m1, m2, m3 = st.columns(3)
                m1.metric(f"Poids moyen {p1}", f"{pf['poids_moyen_a']*100:.0f} %")
                m2.metric(f"Poids moyen {p2}", f"{pf['poids_moyen_b']*100:.0f} %")
                m3.metric("Ratio de couverture moyen", f"{pf['hedge_ratio_moyen']:.3f}",
                          help=f"Couvrir {p1} par {p2} : β = Cov/Var({p2}).")

                st.markdown("**Volatilité annualisée des stratégies :**")
                st.dataframe(
                    pf["resume"].style.format("{:.2f}")
                    .background_gradient(cmap="RdYlGn_r", subset=["Volatilité annualisée (%)"]),
                    width="stretch",
                )
                telecharger_df(pf["resume"], "⬇️ Télécharger le récapitulatif (Excel)",
                               f"portefeuille_resume_{p1}_{p2}.xlsx", key="dl_pf_resume_biv")
                r_a = pf["reduction_vs_a"] * 100
                r_b = pf["reduction_vs_5050"] * 100
                st.success(
                    f"💡 Le portefeuille min-variance réduit la variance de **{r_a:.0f} %** "
                    f"par rapport à un placement 100 % {p1}, et de **{r_b:.0f} %** "
                    f"par rapport à un 50/50. C'est la **prime de diversification** chiffrée."
                )

                s = pf["series"]
                fig, ax = plt.subplots(figsize=(10, 3.2))
                ax.plot(s.index, s[f"Poids {p1}"] * 100, color=BLEU, lw=1, label=p1)
                ax.plot(s.index, s[f"Poids {p2}"] * 100, color=ORANGE, lw=1, label=p2)
                ax.axhline(0, color="0.6", lw=0.6); ax.axhline(100, color="0.6", lw=0.6)
                ax.set_ylabel("Poids (%)"); ax.set_title("Poids optimaux dans le temps")
                ax.legend(loc="upper left", fontsize=8)
                st.pyplot(fig)

                fig, ax = plt.subplots(figsize=(10, 3.2))
                ax.plot(s.index, s[f"Vol. {p1}"], color="0.6", lw=0.7, label=f"100 % {p1}")
                ax.plot(s.index, s[f"Vol. {p2}"], color="0.4", lw=0.7, label=f"100 % {p2}")
                ax.plot(s.index, s["Vol. min-variance"], color=ROUGE, lw=1.4, label="Min-variance (DCC)")
                ax.set_ylabel(f"Volatilité conditionnelle (%/{unite_abbr})")
                ax.set_title("Volatilité : portefeuille optimal vs actifs seuls")
                ax.legend(loc="upper left", fontsize=8)
                st.pyplot(fig)

                telecharger_df(s, "⬇️ Télécharger les poids et volatilités (Excel)",
                               f"portefeuille_{p1}_{p2}.xlsx")
            except Exception as e:
                st.error(f"Erreur : {e}")

        # ==============================================================
        # Cas B — portefeuille multivarié (N actifs)
        # ==============================================================
        else:
            noms_pf = list(dm["noms"])
            st.caption(f"Actifs issus du DCC multivarié : **{', '.join(noms_pf)}** "
                       f"— {dm['n_obs']} {unite_pl}.")
            try:
                pf = ec.dcc_min_variance_portfolio_mv(
                    dm, returns, allow_short=not long_only,
                    periods_per_year=periods_per_year,
                )

                # poids moyens (tableau, car N peut être grand)
                st.markdown("**Poids moyens du portefeuille optimal :**")
                pm = pd.DataFrame(
                    {"Poids moyen (%)": [pf["poids_moyens"][n] * 100 for n in noms_pf]},
                    index=noms_pf,
                )
                st.dataframe(pm.style.format("{:.1f}"), width="stretch")
                telecharger_df(pm, "⬇️ Télécharger les poids moyens (Excel)",
                               "portefeuille_poids_moyens.xlsx", key="dl_pf_poids_mv")

                st.markdown("**Volatilité annualisée des stratégies :**")
                st.dataframe(
                    pf["resume"].style.format("{:.2f}")
                    .background_gradient(cmap="RdYlGn_r", subset=["Volatilité annualisée (%)"]),
                    width="stretch",
                )
                telecharger_df(pf["resume"], "⬇️ Télécharger le récapitulatif (Excel)",
                               "portefeuille_resume_multivarie.xlsx", key="dl_pf_resume_mv")
                r_eq = pf["reduction_vs_eq"] * 100
                meilleur = min(pf["reduction_vs_asset"], key=pf["reduction_vs_asset"].get)
                r_best = pf["reduction_vs_asset"][meilleur] * 100
                st.success(
                    f"💡 Le portefeuille min-variance à {len(noms_pf)} actifs réduit la variance "
                    f"de **{r_eq:.0f} %** par rapport à un équipondéré (1/N), et de "
                    f"**{r_best:.0f} %** par rapport à l'actif le moins risqué ({meilleur}). "
                    f"C'est la **prime de diversification** chiffrée."
                )

                s = pf["series"]
                # poids dans le temps (une courbe par actif)
                fig, ax = plt.subplots(figsize=(10, 3.4))
                couleurs = plt.cm.tab10(np.linspace(0, 1, len(noms_pf)))
                for n, col in zip(noms_pf, couleurs):
                    ax.plot(s.index, s[f"Poids {n}"] * 100, lw=1, color=col, label=n)
                ax.axhline(0, color="0.6", lw=0.6); ax.axhline(100, color="0.6", lw=0.6)
                ax.set_ylabel("Poids (%)"); ax.set_title("Poids optimaux dans le temps")
                ax.legend(loc="upper left", fontsize=8, ncol=min(len(noms_pf), 4))
                st.pyplot(fig)

                # volatilité : min-variance vs équipondéré
                fig, ax = plt.subplots(figsize=(10, 3.2))
                ax.plot(s.index, s["Vol. équipondéré"], color="0.5", lw=0.9,
                        label="Équipondéré (1/N)")
                ax.plot(s.index, s["Vol. min-variance"], color=ROUGE, lw=1.4,
                        label="Min-variance (DCC)")
                ax.set_ylabel(f"Volatilité conditionnelle (%/{unite_abbr})")
                ax.set_title("Volatilité : portefeuille optimal vs équipondéré")
                ax.legend(loc="upper left", fontsize=8)
                st.pyplot(fig)

                telecharger_df(s, "⬇️ Télécharger les poids et volatilités (Excel)",
                               "portefeuille_multivarie.xlsx")
            except Exception as e:
                st.error(f"Erreur : {e}")

# ----------------------------------------------------------------------
# Onglet 9 — Rapport Excel complet (autour d'une série principale)
# ----------------------------------------------------------------------
with onglets[9]:
    st.subheader("Rapport Excel complet")
    st.markdown(
        "Génère **un seul fichier Excel multi-feuilles**, organisé autour d'une "
        "**série principale** (la pierre angulaire) comparée à toutes les autres : "
        "données & rendements, statistiques descriptives, stationnarité (ADF), effets "
        "ARCH-LM, DCC-GARCH, causalité de Granger et Forbes-Rigobon."
    )
    if len(series_cols) < 2:
        st.warning("Il faut au moins **2 séries** pour générer un rapport.")
    else:
        principale = st.selectbox("Série principale (pierre angulaire)",
                                  series_cols, key="rap_principale")
        autres_rap = [c for c in series_cols if c != principale]
        st.caption(f"Comparée à : **{', '.join(autres_rap)}**.")

        inclure_fr_rap = st.checkbox(
            "Inclure le test de contagion Forbes-Rigobon (nécessite une fenêtre de crise)",
            value=True, key="rap_fr",
        )
        crise_d = crise_f = None
        if inclure_fr_rap:
            if not index_dates:
                st.info("Les dates ne sont pas reconnues : Forbes-Rigobon sera ignoré.")
            else:
                dmin_r = returns.index.min().date()
                dmax_r = returns.index.max().date()
                covid_d = max(dmin_r, pd.Timestamp("2020-02-19").date())
                covid_f = min(dmax_r, pd.Timestamp("2020-06-30").date())
                if covid_d >= covid_f:          # données hors COVID → plage complète
                    covid_d, covid_f = dmin_r, dmax_r
                cc1, cc2 = st.columns(2)
                crise_d = cc1.date_input("Début de crise", value=covid_d,
                                         min_value=dmin_r, max_value=dmax_r, key="rap_cd")
                crise_f = cc2.date_input("Fin de crise", value=covid_f,
                                         min_value=dmin_r, max_value=dmax_r, key="rap_cf")

        maxlag_rap = st.slider("Retards maximum (Granger)", 1, 8, 4, key="rap_lag")
        st.caption("Le rapport respecte la **Plage d'étude** et la **fréquence** "
                   "réglées dans la barre latérale.")

        if st.button("📑 Générer le rapport Excel", key="run_rapport"):
            try:
                with st.spinner("Calcul de tous les tests (cela peut prendre ~30 s)…"):
                    feuilles = ec.rapport_complet(
                        returns, principale, prices=prices,
                        crise_debut=crise_d, crise_fin=crise_f,
                        maxlag_granger=maxlag_rap,
                        inclure_fr=inclure_fr_rap and index_dates,
                        freq_label=freq,
                    )
                    xlsx_bytes = ecrire_rapport_xlsx(feuilles)
                st.session_state["rapport_xlsx"] = xlsx_bytes
                st.session_state["rapport_nom"] = f"rapport_{principale}.xlsx"
                st.session_state["rapport_feuilles"] = list(feuilles.keys())
            except Exception as e:
                st.session_state.pop("rapport_xlsx", None)
                st.error(f"Erreur lors de la génération : {e}")

        if st.session_state.get("rapport_xlsx") is not None:
            f_list = st.session_state["rapport_feuilles"]
            st.success(f"✅ Rapport prêt : **{len(f_list)} feuilles** générées.")
            st.download_button(
                "⬇️ Télécharger le rapport Excel",
                st.session_state["rapport_xlsx"],
                file_name=st.session_state["rapport_nom"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_rapport",
            )
            st.caption("Feuilles : " + " · ".join(f_list))

# ----------------------------------------------------------------------
# Onglet 10 — Méthodologie
# ----------------------------------------------------------------------
with onglets[10]:
    st.subheader("Méthodologie et interprétation")
    # --- 0. Préparation des données ------------------------------------
    st.markdown("#### 0. Préparation des données (commune à tous les tests)")
    st.markdown(
        "- La 1ʳᵉ colonne est convertie en **dates** (`pandas.to_datetime`), les lignes triées "
        "par ordre chronologique ; les autres colonnes sont forcées en numérique (virgules "
        "décimales et espaces gérés).\n"
        "- Si vous fournissez des **prix**, l'app calcule les **rendements logarithmiques** "
        "rₜ = ln(Pₜ / Pₜ₋₁) × 100 (`to_log_returns`, en %) ; option rendements arithmétiques. "
        "Des rendements fournis directement sont utilisés tels quels.\n"
        "- Le curseur **« Plage d'étude »** (barre latérale) filtre les dates **avant** tout "
        "calcul : chaque test est estimé uniquement sur la sous-période choisie.\n"
        "- Le **seuil de significativité** (1 %, 5 %, 10 %) pilote tous les verdicts affichés."
    )

    # --- 1. ADF --------------------------------------------------------
    st.markdown("#### 1. Test ADF (Augmented Dickey-Fuller) — stationnarité")
    st.latex(r"\Delta y_t = \alpha + \beta\,t + \gamma\,y_{t-1} "
             r"+ \sum_i \delta_i\,\Delta y_{t-i} + \varepsilon_t")
    st.markdown(
        "**H₀ : γ = 0** (racine unitaire, série non stationnaire). Une p-value faible ⟹ série "
        "**stationnaire**. Les **prix** sont généralement I(1) ; leurs **rendements** sont I(0).\n\n"
        "> **Réalisation :** `statsmodels.tsa.stattools.adfuller`. Le **nombre de retards est "
        "choisi par le critère AIC** (`autolag=\"AIC\"`). Spécification sélectionnable : constante "
        "(`c`), constante + tendance (`ct`, défaut pour les prix) ou aucune (`n`). p-value et "
        "valeurs critiques par la méthode de MacKinnon."
    )

    # --- 2. ARCH-LM ----------------------------------------------------
    st.markdown("#### 2. Test ARCH-LM (Engle, 1982) — effets ARCH")
    st.latex(r"LM = T\cdot R^2 \;\sim\; \chi^2_q "
             r"\quad\text{(régression de } \hat\varepsilon_t^2 \text{ sur ses } q \text{ retards)}")
    st.markdown(
        "**H₀ : pas d'effets ARCH.** Un rejet ⟹ **volatilité groupée** ⟹ un modèle GARCH est "
        "pertinent.\n\n"
        "> **Réalisation :** `statsmodels.stats.diagnostic.het_arch`, sur la série **centrée** "
        "(rendement moins sa moyenne). **Nombre de retards** réglable (curseur). On reporte la "
        "statistique LM et sa p-value χ², ainsi que la version en test F."
    )

    # --- 3. GARCH ------------------------------------------------------
    st.markdown("#### 3. GARCH(p, q) univarié (Bollerslev, 1986)")
    st.latex(r"\sigma_t^2 = \omega + \sum_{i=1}^{p}\alpha_i\,\varepsilon_{t-i}^2 "
             r"+ \sum_{j=1}^{q}\beta_j\,\sigma_{t-j}^2")
    st.markdown(
        "- **α** : réaction aux chocs récents ; **β** : mémoire de la volatilité.\n"
        "- **Persistance = α + β** ; proche de 1 ⟹ chocs durables.\n"
        "- **EGARCH / GJR-GARCH** captent l'**asymétrie** (effet de levier).\n\n"
        "> **Réalisation :** package `arch` (`arch_model`, `rescale=False`), estimé par **maximum "
        "de vraisemblance** (`fit`). Réglables : ordres p, q, équation de moyenne (constante / "
        "zéro / AR), type (GARCH, EGARCH, GJR-GARCH) et **loi des innovations** (normale, "
        "**Student** par défaut, Student asymétrique, GED). La persistance vaut α + β (+ γ/2 si "
        "terme asymétrique) ; AIC/BIC et la demi-vie ln(0,5)/ln(α + β) sont reportés. Un **ARCH-LM "
        "sur les résidus standardisés** vérifie qu'il ne reste pas d'effet ARCH."
    )

    # --- 4. Granger ----------------------------------------------------
    st.markdown("#### 4. Causalité de Granger (1969)")
    st.markdown(
        "X **cause** Y au sens de Granger si les valeurs passées de X améliorent la prévision de "
        "Y. ⚠️ Requiert des séries **stationnaires** (rendements) ; c'est une **précédence "
        "temporelle**, pas une causalité structurelle.\n\n"
        "> **Réalisation :** `statsmodels.tsa.stattools.grangercausalitytests` sur les deux séries "
        "de rendements alignées. Le test est lancé dans les **deux sens** et pour tous les retards "
        "de 1 au maximum choisi ; on reporte la **statistique F** (`ssr_ftest`) et sa p-value par "
        "retard."
    )

    # --- 5. Forbes-Rigobon --------------------------------------------
    st.markdown("#### 5. Contagion de Forbes-Rigobon (2002)")
    st.latex(r"\rho^{*} = \frac{\rho}{\sqrt{1 + \delta\,(1 - \rho^{2})}}, "
             r"\qquad \delta = \frac{\sigma^2_{\text{crise}}}{\sigma^2_{\text{stable}}} - 1")
    st.markdown(
        "où σ² est la variance du **marché-source**. On compare ρ\\* (crise) à ρ (stable).\n"
        "- **H₀ :** ρ\\* = ρ stable ⟹ *pas de contagion*.\n"
        "- **H₁ :** ρ\\* > ρ stable ⟹ *contagion*.\n\n"
        "> **Réalisation :** calcul direct en `numpy`/`scipy`. La **fenêtre de crise** est "
        "délimitée par un curseur de dates (le reste = période stable). Les corrélations "
        "stable/crise sont des Pearson (`np.corrcoef`) sur chaque sous-échantillon ; δ vient des "
        "variances du marché-source. Le test compare ρ\\* et ρ stable via la **transformation z de "
        "Fisher** (SE = √[1/(n_crise − 3) + 1/(n_stable − 3)]), en unilatéral par défaut. Chaque "
        "sous-période doit compter ≥ 5 observations.\n\n"
        "> Conclusion célèbre : « **No Contagion, Only Interdependence** »."
    )

    # --- 6. DCC bivarié ------------------------------------------------
    st.markdown("#### 6. DCC-GARCH bivarié (Engle, 2002)")
    st.latex(r"Q_t = (1 - a - b)\,\bar Q + a\,z_{t-1}z_{t-1}' + b\,Q_{t-1}, "
             r"\qquad R_t = \operatorname{diag}(Q_t)^{-1/2}\,Q_t\,\operatorname{diag}(Q_t)^{-1/2}")
    st.markdown(
        "**a + b < 1** : la corrélation ρₜ = Rₜ[1,2] est stationnaire et revient vers sa "
        "moyenne.\n\n"
        "> **Réalisation (2 étapes, maison) :** (1) un **GARCH univarié** sur chaque série "
        "(package `arch`) donne les **résidus standardisés** zₜ ; (2) la matrice inconditionnelle "
        "Q̄ est la covariance empirique des zₜ, puis **(a, b) est estimé par maximum de "
        "vraisemblance** (log-vraisemblance gaussienne de la corrélation, "
        "`scipy.optimize.minimize`, Nelder-Mead, contrainte a, b > 0 et a + b < 1). Les "
        "**écarts-types** viennent d'une **Hessienne numérique** (différences finies), donc "
        "approximatifs. Sortie : la série ρₜ + moyennes par sous-période."
    )

    # --- 6 bis. DCC multivarié ----------------------------------------
    st.markdown("#### 6 bis. DCC-GARCH multivarié (N séries)")
    st.latex(r"\bar\rho_t = \frac{2}{N(N-1)}\sum_{i<j} R_t[i,j]"
             r"\qquad\text{(indice d'intégration)}")
    st.markdown(
        "Même modèle, généralisé à **N marchés** : Qₜ et Rₜ deviennent des **matrices N × N**, et "
        "les scalaires **(a, b) sont communs à toutes les paires** (DCC « scalaire » d'Engle). "
        "L'indice d'intégration ρ̄ₜ résume la corrélation moyenne entre tous les couples, semaine "
        "par semaine.\n\n"
        "> **Réalisation (`dcc_garch_multivariate`, maison) :** (1) un **GARCH univarié par "
        "série** donne le vecteur des résidus standardisés zₜ ; (2) Q̄ = cov(zₜ), puis (a, b) sont "
        "estimés par **maximum de vraisemblance gaussienne multivariée** "
        "(−½ ∑ₜ [ ln|Rₜ| + zₜ′ Rₜ⁻¹ zₜ ], `scipy.optimize`, Nelder-Mead). À chaque date, ln|Rₜ| "
        "et zₜ′ Rₜ⁻¹ zₜ utilisent `numpy.linalg.slogdet` et `solve` ; écarts-types par **Hessienne "
        "numérique**. Sorties : **matrice de corrélation moyenne** R̄, indice d'intégration ρ̄ₜ, et "
        "la corrélation conditionnelle de **chaque couple** (i, j)."
    )

    # --- 7. Diebold-Yilmaz --------------------------------------------
    st.markdown("#### 7. Spillover de volatilité — Diebold & Yilmaz (2012)")
    st.latex(r"S = \frac{1}{N}\sum_{i\neq j}\tilde\theta_{ij}\times 100"
             r"\qquad \text{Net}_j = (\text{émis}) - (\text{reçu})")
    st.markdown(
        "Décomposition **généralisée** (KPPS, invariante à l'ordre) de la variance d'erreur de "
        "prévision d'un VAR à l'horizon H. θ̃ᵢⱼ = part de la volatilité future de i due à j.\n\n"
        "> **Réalisation :** les **volatilités conditionnelles** de chaque série viennent d'un "
        "GARCH univarié (`garch_vol_matrix`). Un **VAR** est estimé dessus "
        "(`statsmodels.tsa.api.VAR`) ; la décomposition généralisée se calcule à partir de la "
        "matrice de covariance des résidus (Σ) et des matrices moyennes mobiles (`ma_rep`), puis "
        "**chaque ligne est normalisée à 100 %**. Retards du VAR, horizon H, loi des marges et "
        "**fenêtre glissante** sont réglables ; l'indice glissant ré-estime le modèle sur des "
        "fenêtres successives. Calculs **mis en cache**."
    )

    # --- 8. Portefeuille ----------------------------------------------
    st.markdown("#### 8. Portefeuille à variance minimale dynamique (covariance DCC)")
    st.latex(r"H_t = D_t R_t D_t, \qquad w_t = \frac{H_t^{-1}\mathbf 1}{\mathbf 1'\,H_t^{-1}\mathbf 1}, "
             r"\qquad w_{1,t} = \frac{\sigma_{2,t}^2 - \operatorname{cov}_t}"
             r"{\sigma_{1,t}^2 + \sigma_{2,t}^2 - 2\,\operatorname{cov}_t}")
    st.markdown(
        "> **Réalisation :** la covariance conditionnelle est reconstruite à partir du **DCC "
        "estimé à l'onglet 6** (ρₜ et les volatilités GARCH des deux marges). Les poids sont "
        "calculés à chaque période (option **sans vente à découvert** : poids bornés à [0, 1]). "
        "Les **rendements de portefeuille** utilisent les **poids décalés d'une période** (pas "
        "d'anticipation). On compare la **volatilité annualisée** (× √52 en hebdomadaire, × √252 "
        "en quotidien — selon la fréquence choisie dans la barre latérale) à un 50/50 et aux actifs "
        "seuls ; la **réduction de variance** chiffre la prime de diversification. Ratio de "
        "couverture βₜ = covₜ / σ²₂,ₜ."
    )

    st.markdown("---")
    st.markdown("**Récapitulatif des outils** (code dans `econometrics.py`) :")
    st.markdown(
        "| Test | Bibliothèque / fonction | Estimation |\n"
        "|------|------------------------|-----------|\n"
        "| ADF | `statsmodels.adfuller` | retards par AIC |\n"
        "| ARCH-LM | `statsmodels.het_arch` | régression auxiliaire |\n"
        "| GARCH | `arch.arch_model` | max. de vraisemblance |\n"
        "| Granger | `statsmodels.grangercausalitytests` | test F sur VAR |\n"
        "| Forbes-Rigobon | `numpy` / `scipy` | corrélations + z de Fisher |\n"
        "| DCC-GARCH bivarié | maison + `arch` + `scipy.optimize` | 2 étapes, MLE |\n"
        "| DCC-GARCH multivarié | maison + `arch` + `scipy.optimize` | Rₜ N×N, MLE gaussienne |\n"
        "| Diebold-Yilmaz | `statsmodels.VAR` + maison | GFEVD généralisée |\n"
        "| Portefeuille | maison (`numpy`) | min-variance analytique |"
    )

st.divider()
st.caption(
    "Application d'économétrie financière • ADF · ARCH-LM · GARCH · Granger · "
    "Forbes-Rigobon · DCC-GARCH · Diebold-Yilmaz · Portefeuille • construite avec Streamlit "
    "• © Ali EB"
)
